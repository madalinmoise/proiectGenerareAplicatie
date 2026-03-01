from openpyxl import load_workbook
import os

data_file = "temp_uploads/fisier_excel.xlsx"
try:
    wb = load_workbook(data_file, read_only=True)
    print(f"Sheets: {wb.sheetnames}")
    ws = wb.active
    print(f"Active Sheet: {ws.title}")
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    print(f"Header: {header}")
    count = 0
    for r in rows:
        count += 1
        if count <= 2:
            print(f"Row {count}: {r}")
    print(f"Total rows found by openpyxl: {count}")
except Exception as e:
    print(f"Error: {e}")
